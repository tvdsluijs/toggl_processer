#https://pypi.org/project/TogglPy/
import time
from datetime import datetime, date, timedelta
import json
import math

from toggl.TogglPy import Toggl
from openpyxl import load_workbook

from functions.readConfig import readConfig


class TogglMe:
    def __init__(self, user_agent=None, workspace_id=None, start_col=0, start_row=0, name=""):
        if user_agent == workspace_id is None:
            return None

        cf = readConfig()
        self.config = cf.config

        self.start_col = start_col
        self.start_row = start_row

        self.months = ["Unknown",
                  "January",
                  "Febuary",
                  "March",
                  "April",
                  "May",
                  "June",
                  "July",
                  "August",
                  "September",
                  "October",
                  "November",
                  "December"]

        self.normal_work_hours = 7

        self.user_agent = user_agent
        self.user_name = name
        self.workspace_id = workspace_id
        self.response = None
        self.last_month_first = None
        self.last_month_last = None
        self.curr_year = None
        self.filename = None

        self.my_excel_data = []
        self.excel_template = "excel/urenoverzicht.xlsx"
        self.excel_file = None

        self.about_last_month()
        self.get_data()
        self.parse_data()
        self.create_excel()

    def about_last_month(self):
        now = time.localtime()
        last_month = now.tm_mon - 1 if now.tm_mon > 1 else 12
        year = now.tm_year if now.tm_mon > 1 else now.tm_year - 1
        self.curr_year = year
        self.filename = "data/{}{}".format(year, last_month)
        self.excel_file = "output/{}-{}-{}-{}.xlsx".format(self.user_name.replace(' ', '-').lower(), "hours", year,
                                                           self.months[last_month].lower())

        # Get the last day of last month by taking the first day of this month
        # and subtracting 1 day.
        last = date(now.tm_year, now.tm_mon, 1) - timedelta(1)

        # Set the day to 1 gives us the start of last month
        first = last.replace(day=1)

        # The default string representation of these datetime instances is
        # YYYY-mm-dd format (which is what I usually need),
        # so we can just print them out
        self.last_month_first = first
        self.last_month_last = last

    def get_data(self):
        toggl = Toggl()
        toggl.setAPIKey(self.config['api'])

        data = {
            'workspace_id': workspace_id,
            'user_agent': user_agent,
            'since': self.last_month_first,
            'until': self.last_month_last
        }
        self.response = toggl.request("https://toggl.com/reports/api/v2/details", parameters=data)
        self.save_raw_data()

    def parse_data(self):
        data = {}
        for d in self.response['data']:
            date = self.convert_timedate(d['start'])
            try:
                data[date[1]] += d['dur']
            except KeyError:
                data[date[1]] = d['dur']

        for k, v in data.items():
            toggl_date = self.convert_strdate(k)
            toggl_time = self.convert_millis(v)
            round_quarter = self.round_quarter(toggl_time[2])
            worked_overtime = self.minus_normal_hours(round_quarter)
            row = toggl_date.day + self.start_row
            col = toggl_date.month + self.start_col
            data = col, row, worked_overtime
            self.my_excel_data.append(data)

    def create_excel(self):
        wb = load_workbook(self.excel_template)

        ws = wb.active

        ws.cell(row=1, column=2, value=self.user_name) # name field in excel
        ws.cell(row=1, column=5, value=self.curr_year) # year field in excel

        for data in self.my_excel_data:
            ws.cell(row=data[1], column=data[0], value=data[2])

        wb.save(filename=self.excel_file)

    @staticmethod
    def convert_timedate(date_time_str=None):
        date_time_obj = datetime.strptime(date_time_str, '%Y-%m-%dT%H:%M:%S+02:00')
        return date_time_obj, date_time_obj.strftime("%Y-%m-%d")

    @staticmethod
    def convert_strdate(date_str=None):
        date_time_obj = datetime.strptime(date_str, '%Y-%m-%d')
        return date_time_obj

    @staticmethod
    def round_half_up(n, decimals=0):
        multiplier = 10 ** decimals
        return math.floor(n * multiplier + 0.5) / multiplier

    def minus_normal_hours(self, hours=0):
        if hours > self.normal_work_hours:
            return hours-self.normal_work_hours
        return hours

    @staticmethod
    def round_quarter(x):
        return round(x * 4) / 4

    def convert_millis(self, milli=None):
        seconds = self.round_half_up((milli / 1000) % 60, 2)
        minutes = self.round_half_up((milli / (1000 * 60)) % 60, 2)
        hours = self.round_half_up(((milli / (1000 * 60 * 60)) % 24), 2)
        return seconds, minutes, hours

    def save_raw_data(self):
        app_json = json.dumps(self.response)

        f = open("{}.json".format(self.filename), "w")
        f.write(app_json)
        f.close()


if __name__ == '__main__':
    user_agent = "theo@vandersluijs.nl"
    workspace_id = 881607
    name = "Theo van der Sluijs"
    t = TogglMe(user_agent, workspace_id, 1, 3, name)

# todayDate = datetime.date.today()
# if todayDate.day > 25:
#     todayDate += datetime.timedelta(7)
# print(todayDate.replace(day=1))

# response = toggl.request("https://toggl.com/reports/api/v2/weekly")

# data = {
#          'workspace_id': workspace_id,
#          'user_agent': user_agent,
#          'since':'2019-06-01',
#          'until':'2019-06-30'
#         }
# response = toggl.request("https://toggl.com/reports/api/v2/summary", parameters=data)

# data = {
#          'workspace_id': workspace_id,
#          'user_agent': user_agent
#         }
# response = toggl.request("https://toggl.com/reports/api/v2/weekly", parameters=data)


# for line in response:
#     print(line)


# print(response)

# print(toggl.getWorkspaces())
# print(toggl.getClients())



